from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from xlsxwriter.utility import xl_col_to_name
import re
import os

def run():
    wb = Workbook()
    ws = wb.active
    BoldFont = Font(bold=True)
    Void = PatternFill(start_color="888888", end_color="888888", bgColor="888888", fill_type="solid")
    Gray = PatternFill(start_color="D8D8D8", end_color="D8D8D8", bgColor="D8D8D8", fill_type="solid")
    Good = PatternFill(start_color="C6EFCE", end_color="C6EFCE", bgColor="C6EFCE", fill_type="solid")

    # Column and Row setup
    idxs = {}
    Brands = list(map(lambda x: x[:-4].replace("-", " "), os.listdir("../../Links")))
    Brands = [""] + Brands

    for b in range(len(Brands)):
        idxs[Brands[b].replace("-", "").replace(" ", "")] = xl_col_to_name(b)

    with open("../../Docs/Base.csv", "r") as f:
        Specs = list(map(lambda x: x.rstrip(), f.read().split(',')))

    Specs += ["_"] # the defualt case

    ws.append(Brands)
    for i in range(len(Specs)):
        ws["A"+str(i+2)] = Specs[i]
        idxs[Specs[i]] = str(i+2)

    for cell in ws["A"]:
        cell.font = BoldFont
    for cell in ws[1]:
        cell.font = BoldFont
        cell.alignment = Alignment(horizontal="center", vertical="center")


    # Fill in the cells
    files = os.listdir("Corrections")
    files.remove("__init__.py")
    files.remove("__pycache__")
    files = list(map(lambda x: x.replace(".py", ""), files))
    
    def filt(x):
        return re.search(r"^\t{4}case", x) != None

    for f in files:
        brand = f.replace(".py", "")
        with open("Corrections/" + f + ".py", "r") as file:
            data = file.readlines()
        
        idx = list(map(lambda x: data.index(x), filter(filt, data)))
    
        for i in idx:
            s = re.search(r'(".*"|_)', data[i])
            if s is not None:
                spec = data[i][s.span()[0]:s.span()[1]].replace('"', '')

                if "pass" in data[i+1]:
                    ws[idxs[brand]+idxs[spec]].fill = Gray
                else:
                    ws[idxs[brand]+idxs[spec]].fill = Good
            else:
                # print(data[i])
                pass #! default case is not getting handled

    # black out brands that do not have any corrections (they have no data to being with)
    bad_brands = set(map(lambda x: x.replace(" ", ""), Brands)).difference(set(files)).difference(set(['']))
    for bb in bad_brands:
        for s in Specs:
            ws[idxs[bb]+idxs[s]].fill = Void

    # Iterate over all columns and adjust their widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length+2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save("../../Docs/CorrectionStatus.xlsx")


if __name__ == "__main__":
    run()